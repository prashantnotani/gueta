from django.http import HttpResponse
from django.shortcuts import render
import mysql.connector
from mysql.connector import Error
from django.core.files.storage import FileSystemStorage
from django.core.mail import EmailMessage
from django.conf import settings
import openpyxl
from openpyxl import load_workbook
from django.shortcuts import redirect
from django.contrib import messages
# Create your views here.
def index(requests):
    return render(requests,"index.html")

def login(request):
    return render(request,"login.html")

def facreg(request):
    try:

        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        query = "select * from city"
        cursor.execute(query)
        city = cursor.fetchall()
        return render(request,"facultyreg.html", {'city':city})
    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()

def getarea(request):
    try:

        cid = int(request.GET.get('cid'))
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        query = "select * from area where cid = '%d'" %cid
        cursor.execute(query)
        area = cursor.fetchall()

        return render(request, 'area.html',{'area':area})

    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()


def savedata(request):
    try:
        firstName = request.GET.get("firstName")
        middleName= request.GET.get("middleName")
        lastName = request.GET.get("lastName")
        raddress = request.GET.get("raddress")
        city = int(request.GET.get("city"))
        area = int(request.GET.get("area"))
        email = request.GET.get("email")
        mobile = int(request.GET.get("mobile"))
        password = request.GET.get("password")
        designation = request.GET.get("designation")
        cname = request.GET.get("collegeName")
        caddress = request.GET.get("caddress")
        ccity = int(request.GET.get("ccity"))
        carea = int(request.GET.get("carea"))
        cname = request.GET.get("collegeName")
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor(buffered=True)

        query3 = "insert into profreg(firstName,middleName,lastName,email,password,mobile,designation,cname) VALUES ('%s','%s','%s','%s','%s','%s','%d','%s','%s')"%(firstName,middleName,lastName,email,password,cpassword,mobile,designation,cname)
        cursor.execute(query3)
        query4 = "select prid from profreg WHERE mobile=('%d')"%(mobile)
        cursor.execute(query4)
        rec = cursor.fetchone()

        query1 = "insert into address(raddress,areaid,prid) VALUES ('%s','%d','%d')"%(raddress,area,rec[0])
        query2 = "insert into address(caddress,areaid,prid) VALUES ('%s','%d','%d')"%(caddress,carea,rec[0])
        cursor.execute(query2)
        cursor.execute(query1)
        conn.commit()
        return HttpResponse(login(request))
    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()

def logindata(request):
    try:
        print("<script>alert('sjafnk');</script>")
        email = request.GET.get("mail")
        password=request.GET.get("password")
        designation = request.GET.get("designation")
        gmail = request.GET.get('e')
        print(gmail,password,designation)
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        if(gmail):
            query="select firstName from profreg where email = '%s'"%(gmail)
        else:
            query = "select firstName from profreg where email = '%s' AND password = '%s' AND designation='%s'"%(email,password,designation)
        cursor.execute(query)

        #cursor.execute(query)
        rec = cursor.fetchone()
        if rec == None:
            print("**********************")
            return render(request,"out.html")
        else:
            lname=rec[0]
            conn.commit()
            request.session['lname']=lname
            if(request.session.get('lname')):
                return redirect('index')
            else:
                print("error")
    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()


def about(request):
    return render(request,"about.html")

def signout(request):
    del request.session['lname']
    return redirect('index')

def upload(request):
    if(request.session.get('lname')):
        return render(request,"upload.html")
    else:
        return render(request,"login.html")

def daol(request):
    try:
       conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
       cursor = conn.cursor()
       if request.method == 'POST' and request.FILES['myfile']:
            myfile = request.FILES['myfile']
            name = request.POST.get('name')
            fs = FileSystemStorage()
            filename= myfile.name
            extension = filename.split('.')
            uploaded_file_name = name + "." +extension[1]
            filename = fs.save(uploaded_file_name, myfile)
            uploaded_file_url = fs.url(filename)
            query = "insert into uploadr(researchpaper,name1) VALUES ('%s','%s')" %(uploaded_file_url,name)
            cursor.execute(query)
            conn.commit()
            return HttpResponse(index(request))

    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()



def adde(request):
    return render(request,"adde.html")

def eins(request):
    try:
        eventName=request.GET.get("eventName")
        eventDate=request.GET.get("eventDate")
        eventDescription = request.GET.get("eventDescription")

        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        query = "insert into event(event_name,event_description,event_date) VALUES ('%s','%s','%s')"%(eventName,eventDescription,eventDate)
        cursor.execute(query)
        conn.commit()
        return HttpResponse(adde(request))
    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()


def event(request):
    try:
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        query = "SELECT * FROM event ORDER BY event_id DESC LIMIT 3"
        cursor.execute(query)
        rec = cursor.fetchall()

        conn.commit()
        return render(request,"events.html",{'rec':rec})
    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()

def viewmore(request):
    try:
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        query = "SELECT * FROM event "
        cursor.execute(query)
        rec = cursor.fetchall()

        conn.commit()
        return render(request,"viewmoreevents.html",{'rec':rec})
    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()


def notification(request):
    return render(request, "notification.html")

def disnot(request):
     try:
        notificationTitle=request.GET.get("notificationTitle")
        notificationMessage = request.GET.get("notificationMessage")
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        query = "insert into notification(name1,message) VALUES ('%s','%s') "%(notificationTitle,notificationMessage)
        email = EmailMessage(notificationTitle, notificationMessage, to=['prashantnotani@gmail.com'])
        email.send()
        cursor.execute(query)
        conn.commit()
        messages.success(request, 'Notification Sent')
        return redirect('notification')

     except Error as e:
        print(e)
     finally:

        cursor.close()
        conn.close()

def searche(request):
    try:
        params=request.GET.get("param")
        print("**********",params)
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        query = "select * from event WHERE event_name LIKE ('%s')"%("%" + params+ "%",)
        cursor.execute(query)
        rec = cursor.fetchall()
        print("**********",rec)
        conn.commit()

        return render(request,"events.html",{'rec':rec})
    except Error as e:
        print(e)
    finally:

        cursor.close()
        conn.close()

def ventreg(request):
      try:

        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        query = "select * from city"
        cursor.execute(query)
        city = cursor.fetchall()
        return render(request,"ventreg.html", {'city':city})
      except Error as e:
        print(e)
      finally:
        cursor.close()
        conn.close()

def avedata(request):
     try:
        firstName = request.GET.get("firstName")
        middleName= request.GET.get("middleName")
        lastName = request.GET.get("lastName")
        raddress = request.GET.get("raddress")
        city = int(request.GET.get("city"))
        area = int(request.GET.get("area"))
        email = request.GET.get("email")
        mobile = int(request.GET.get("mobile"))

        designation = request.GET.get("designation")
        cname = request.GET.get("collegeName")
        caddress = request.GET.get("caddress")
        ccity = int(request.GET.get("ccity"))
        carea = int(request.GET.get("carea"))
        cname = request.GET.get("collegeName")
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor(buffered=True)

        query3 = "insert into eventreg(firstName,middleName,lastName,email,mobile,designation,cname) VALUES ('%s','%s','%s','%s','%d','%s','%s')"%(firstName,middleName,lastName,email,mobile,designation,cname)
        cursor.execute(query3)
        query4 = "select id from eventreg WHERE mobile=('%d')"%(mobile)
        cursor.execute(query4)
        rec = cursor.fetchone()

        query1 = "insert into address(raddress,areaid,prid) VALUES ('%s','%d','%d')"%(raddress,area,rec[0])
        query2 = "insert into address(caddress,areaid,prid) VALUES ('%s','%d','%d')"%(caddress,carea,rec[0])
        cursor.execute(query2)
        cursor.execute(query1)
        conn.commit()
        return HttpResponse(index(request))
     except Error as e:
        print(e)
     finally:
        cursor.close()
        conn.close()

def importfile(request):
    if(request.session.get('lname')):
        return render(request,"importfile.html")
    else:
        return render(request,"login.html")

def contact(request):
    try:
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor(buffered=True)
        if request.method == 'POST' and request.FILES['myfile']:
            myfile = request.FILES['myfile']
            wb = load_workbook(myfile)
            sheet=wb.active

            max_row=sheet.max_row

            max_column=sheet.max_column
            row_data = list()
            for i in range(2,max_row+1):

                # iterate over all columns
                for j in range(1,max_column+1):
                    #  get particular cell value
                    cell_obj=sheet.cell(row=i,column=j)
                    row_data.append(cell_obj.value)
                print(row_data[0])
                query = "insert into profreg(firstName,email) VALUES ('%s','%s')"%(row_data[0],row_data[1])
                cursor.execute(query)
                row_data.clear()
                print('\n')
        conn.commit()
        return HttpResponse(index(request))
    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()

def download(request):

    try:
        conn = mysql.connector.connect(host='localhost', database='gueta', user='root', password='root', port='3307')
        cursor = conn.cursor()
        query = "select * from uploadr"
        cursor.execute(query)
        rec = cursor.fetchall()
        print(rec)
        conn.commit()
        return render(request,"download.html",{'rec':rec})
    except Error as e:
        print(e)
    finally:
        cursor.close()
        conn.close()

